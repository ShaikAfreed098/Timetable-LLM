"""
Export timetable to PDF or Excel.
"""
from __future__ import annotations

import io
from typing import List

from sqlalchemy.orm import Session

from app.core.scheduler import DAYS, TEACHING_PERIODS, PERIOD_TIMES
from app.models.timetable import TimetableSlot
from app.models.batch import Batch


def _get_cell_label(slot: TimetableSlot, db: Session) -> str:
    if slot.slot_type != "class" or slot.subject_id is None:
        return ""
    from app.models.subject import Subject
    from app.models.faculty import Faculty

    subj = db.get(Subject, slot.subject_id)
    fac = db.get(Faculty, slot.faculty_id) if slot.faculty_id else None
    parts = [subj.code if subj else "?"]
    if fac:
        parts.append(fac.name.split()[-1])  # last name
    return "\n".join(parts)


def export_to_excel(db: Session, timetable_id: str) -> bytes:
    """Return Excel file as bytes."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    slots: List[TimetableSlot] = (
        db.query(TimetableSlot)
        .filter(TimetableSlot.timetable_id == timetable_id)
        .all()
    )
    if not slots:
        raise ValueError("No timetable found for given ID.")

    batch_ids = list({s.batch_id for s in slots})
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for b_id in batch_ids:
        batch = db.get(Batch, b_id)
        sheet_name = batch.name if batch else str(b_id)
        ws = wb.create_sheet(title=sheet_name[:31])

        # Header row: Day | P1 | P2 | ... | P7
        ws.cell(row=1, column=1, value="Day / Period").font = header_font
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).alignment = center

        for p_idx, period in enumerate(TEACHING_PERIODS, start=2):
            cell = ws.cell(row=1, column=p_idx, value=f"P{period}\n{PERIOD_TIMES[period]}")
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            ws.column_dimensions[get_column_letter(p_idx)].width = 18

        ws.column_dimensions["A"].width = 14

        # Data rows
        slot_index = {
            (s.day_of_week, s.period_number): s
            for s in slots
            if s.batch_id == b_id
        }

        for d_idx, day in enumerate(DAYS, start=2):
            ws.cell(row=d_idx, column=1, value=day).alignment = center
            for p_idx, period in enumerate(TEACHING_PERIODS, start=2):
                s = slot_index.get((day, period))
                label = _get_cell_label(s, db) if s else ""
                cell = ws.cell(row=d_idx, column=p_idx, value=label)
                cell.alignment = center

        ws.row_dimensions[1].height = 40

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_to_pdf(db: Session, timetable_id: str) -> bytes:
    """Return PDF file as bytes."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm

    slots: List[TimetableSlot] = (
        db.query(TimetableSlot)
        .filter(TimetableSlot.timetable_id == timetable_id)
        .all()
    )
    if not slots:
        raise ValueError("No timetable found for given ID.")

    batch_ids = list({s.batch_id for s in slots})
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=1 * cm, bottomMargin=1 * cm)
    styles = getSampleStyleSheet()
    story = []

    for b_id in batch_ids:
        batch = db.get(Batch, b_id)
        batch_name = batch.name if batch else str(b_id)

        story.append(Paragraph(f"Timetable — {batch_name}", styles["Heading2"]))
        story.append(Spacer(1, 0.3 * cm))

        slot_index = {
            (s.day_of_week, s.period_number): s
            for s in slots
            if s.batch_id == b_id
        }

        header = ["Day"] + [f"P{p}\n{PERIOD_TIMES[p]}" for p in TEACHING_PERIODS]
        data = [header]
        for day in DAYS:
            row = [day]
            for period in TEACHING_PERIODS:
                s = slot_index.get((day, period))
                row.append(_get_cell_label(s, db) if s else "")
            data.append(row)

        col_widths = [2.5 * cm] + [3.2 * cm] * len(TEACHING_PERIODS)
        table = Table(data, colWidths=col_widths, rowHeights=[1.2 * cm] + [1.5 * cm] * len(DAYS))
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                ]
            )
        )
        story.append(table)
        story.append(Spacer(1, 0.8 * cm))

    doc.build(story)
    return buf.getvalue()
